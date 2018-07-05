import re
import urllib.request
import urllib.parse
import excel_read
from openpyxl import Workbook
productid=excel_read.product_check()
#print(productid)
#productid=['73983701']
#productid=['1232434354','31664301']
import datetime
size=['tiny baby','new baby','Up to 1 mnth','Up to 3 mnths','0-6 months','6-12 months','1-2 years','2-4 years','4-7 years','3-6 months','6-9 months','9-12 months','12-18 months','18-24 months','2-3 years','3-4 years','4-5 years','5-6 years','6-7 years','7-8 years']
size1=['0-6 months','6-12 months','1-2 years','2-4 years','4-7 years','3-6 months','6-9 months','9-12 months','12-18 months','18-24 months','2-3 years','3-4 years','4-5 years','5-6 years','6-7 years','7-8 years']
url = 'https://www.johnlewis.com/search/'
##productid=['MF703','MF773','G0674']
import excel_write
outputwb=excel_write.create_output()
output=outputwb[datetime.date.today().strftime("%Y-%m-%d")]
sale_counter=0
no_availablity_counter=0
#def salecheck(stock_lists):
#    for sale_item in stock_lists:
#            if 'in stock' in sale_item :
#                for size_list in size1:
#                    if size_list in sale_item:
#                        ava=1
#                        return ava
#            else:
#                ava=0
#    return ava

for i in range(0,len(productid)):
  #print('Checking Product '+productid[i])
  try:
    values = {'Ntt':productid[i]}
    print('checking product '+str(productid[i]))
    #print(values)
    data = urllib.parse.urlencode(values)
    #print(data)
    data = data.encode('utf-8')
    #print(data)
    full_url = url + '?Ntt=' + str(productid[i])
    #print(full_url)
    req=urllib.request.Request(full_url)
    #req = urllib.request.Request(url,data)
    #url='https://www.johnlewis.com/search?Ntt=73983701'
    #req = urllib.request.Request(url)
    #print(req)
    resp = urllib.request.urlopen(req)
    respData = resp.read().decode('utf-8')
    #f=open('temp1','w')
    #f.write(respData)
    #f.close()
    result=[]
    stock_lists=[]
    ava=1
    Reduce_to_clear=0
    salecount=0
    line = respData.splitlines()
    #f=open('temp3','w')
    productidtemp=productid[i]

    for i in range (1,len(line)):
        #f.write(line[i]+'\n')
        if '''Sorry, we couldn&#039;t find any results matching ''' in line[i]:
            ava=0
            #print('not available')
            break
        if 'Reduced to clear' in line[i]:
              #salecount=salecount+1
              Reduce_to_clear=1
              #print('Reduce to clear')
        if 'data-jl-size' in line[i]:
            size=re.findall(r'"([^"]*)"', line[i])
            if 'out-of-stock' not in line[i-4]:
                stock_list=str(size[0])+' in stock\n'
            else:

                stock_list=str(size[0])+' out of stock\n'
            stock_lists.append(stock_list+'\n')
            #print(stock_lists)

        #if any(word in line[i] for word in size):
        #    if len(line[i])<30:
        #        if 'out of stock' not in line[i]:
        #            stock_list=str(line[i]+' in stock\n')
        #        else:
        #            stock_list=str(line[i])
        #    stock_lists.append(stock_list+'\n')
    if ava==0:
        no_availablity_counter=no_availablity_counter+1
        row_max=output.max_row
        output.cell(row=row_max+1, column=1, value=productidtemp)
        output.cell(row=row_max+1, column=2, value='N')
        #print('==============================================')
        #print('Product Code: '+values['q']+'\n')
        #print('No product available')
    if Reduce_to_clear==1:
        sale_counter=sale_counter+1
        #no_availablity_counter=no_availablity_counter+1
        row_max=output.max_row
        output.cell(row=row_max+1, column=1, value=productidtemp)
        output.cell(row=row_max+1, column=2, value='N/A')
        output.cell(row=row_max+1, column=3, value='Y')
        output.cell(row=row_max+1, column=4, value=str.join('.\n\n', stock_lists))
        #ava=salecheck(stock_lists)
        #for sale_item in stock_lists:
        #    if 'in stock' in sale_item :
        #        for size_list in size1:
        #            if size_list in sale_item:
        #                ava=1
        #                break
        #    else:
        #        ava=0
        #        continue
        #    break
        #if ava==1:
        #    row_max=output.max_row
        #    output.cell(row=row_max+1, column=1, value=values['q'])
        #    output.cell(row=row_max+1, column=2, value='Y')
        #    output.cell(row=row_max+1, column=3, value='Y')
        #    output.cell(row=row_max+1, column=4, value=str.join('.\n', stock_lists))
        #if ava==0:
        #    no_availablity_counter=no_availablity_counter+1
        #    row_max=output.max_row
        #    output.cell(row=row_max+1, column=1, value=values['q'])
        #    output.cell(row=row_max+1, column=2, value='N')
        #    output.cell(row=row_max+1, column=3, value='Y')
        #   output.cell(row=row_max+1, column=4, value=str.join('.\n', stock_lists))
    #f.close()

        #print('==============================================')
        #print('Product Code: '+values['q']+'\n')
        #print('Product On Sale')
        #for i in range(1,len(stock_lists)):
        #    print(stock_lists[i]+'\n')
  except  urllib.error.URLError as e:
    print(productid[i]+'\n')
    print(e.reason)
  outputwb.save('JohnLewis'+datetime.date.today().strftime("%Y-%m-%d")+'.xlsx')
print('sale count = ',sale_counter)
print('no availablity count = ', no_availablity_counter)
print('Job Done')






